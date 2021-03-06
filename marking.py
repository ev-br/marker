from __future__ import division, print_function, absolute_import

from subprocess import Popen, PIPE, STDOUT, TimeoutExpired
import logging
import datetime
import sys
import os
import contextlib
import argparse
import py_compile
from io import StringIO

import openpyxl

from LMSzip import fill_cohort, Student


@contextlib.contextmanager
def use_folder(folder):
    """Step into an existing folder, then step back."""
    start_folder = os.getcwd()
    os.chdir(folder)
    yield
    os.chdir(start_folder)


def name_from_path(path):
    # XXX: check on Windows?
    if path.endswith('/'):
        path = path[:-1]
    head, tail = os.path.split(path)
    return tail.replace(' ', '_')


def setup_logger(logger_name, log_file, level=logging.INFO, log_buffer=None):
    # http://stackoverflow.com/questions/17035077/python-logging-to-multiple-log-files-from-different-classes
    l = logging.getLogger(logger_name)
    formatter = logging.Formatter('%(levelname)s: %(asctime)s : %(message)s')
    fileHandler = logging.FileHandler(log_file, mode='w')
    fileHandler.setFormatter(formatter)
    streamHandler = logging.StreamHandler()
    streamHandler.setFormatter(formatter)

    l.setLevel(level)
    l.addHandler(fileHandler)
    l.addHandler(streamHandler)
    if log_buffer is not None:
        captureHandler = logging.StreamHandler(log_buffer)
        l.addHandler(captureHandler)
    return l


class Program(object):
    def __init__(self, folder, fname, logger, timeout=None, *args, **kwds):
        super(Program, self).__init__(*args, **kwds)

        self.workdir = os.path.abspath(folder)
        self.fname = fname

        self.timeout = timeout if timeout else 5
        self.cmd = None

    def compile(self, logger):
        """ Compile the code. Return True/False for success status.
        """
        # XXX: refactor the python-specific part to _compile.
        logger.info("Compiling %s ..." % self.fname)

        self.cmd = [sys.executable, self.fname]
        with use_folder(self.workdir):
            try:
                # Check if the file is valid python code.
                py_compile.compile(os.path.join(self.workdir, self.fname),
                                   doraise=True)
                success = True
            except Exception as e:
                logger.error("Compilation failed. Exception %s " % e)
                success = False
        return success

    def run(self, logger, inp=None):
        """Run the program in a subprocess. Grab the output.
        """
        logger.info("running %s with input %s" % (self.fname, inp))
        inp_ = str(inp) if inp is not None else ""
        with use_folder(self.workdir):
            try:
                p = Popen(self.cmd, stdin=PIPE, stdout=PIPE, stderr=PIPE,
                          universal_newlines=True)
                output, err = p.communicate(input=inp_, timeout=self.timeout)
            except TimeoutExpired:
                logger.error("Timed out %s seconds" % self.timeout)
                p.kill()
                output, err = "", None
        return output, err


class FakeProgram(object):
    """Wrap a callable `func` into a Program-compatible interface."""
    def __init__(self, func, timeout):
        self.func = func
        self.timeout = timeout

    def compile(self, logger):
        return True

    def run(self, logger, inp=None):
        logger.debug("calling %s with input %s" % (self.func, inp))
        if inp is None:
            inp = {}
        try:
            res = self.func(**inp)
        except TypeError:
            # inp is not a dict, try a positional argument
            res = self.func(inp)
        return res, None


class Submission(object):
    """A student's submission. Given a folder, find an executable file.

    Attributes
    ----------
    folder : str
        The absolute path to the submission folder
    fname : str
        The name of the executable, relative to self.folder
    """
    def __init__(self, folder, kind='python', *args, **kwds):
        super(Submission, self).__init__(*args, **kwds)
        if not os.path.isdir(folder):
            raise ValueError("Expect a directory, got %s." % folder)
        self.folder = os.path.abspath(folder)

        # find the exectutable
        self.fname = None
        for f in os.listdir(self.folder):
            if self._is_executable(f):
                self.fname = f
                break
        else:
            # did not find it
            raise ValueError("Failed to find an executable in %s." % folder)

    def _is_executable(self, f):
        """Replace for non-python executables."""
        return f.endswith('.py')


class Exercise(object):
    """An Exercise, a list of tasks with their weights.

    Each task is defined by an input and a weighting. `self.inputs` is a list
    of inputs for each task. `self.weights` gives weighting of each task, plus
    a possible mark allocation for the program to compile, out of 100 in total.
    That is, ``len(self.weights) == len(self.inputs) + 1``, with ``weights[0]``
    being the mark allocation for compilation. Total weight is 100.

    To construct an Exercise, construct an instance of `Exercise` with
    either a `base_program` (which can be either a path to a file or a Program
    instance), or a callable with the signature ``func(input)``.

    By default, the output is checked by literal comparison of the output with
    the ground truth. If this is not desired, subclass Exercise and redefine
    the `_check` method.

    To mark an Exercise use the `mark` method which returns the numeric mark.

    """
    def __init__(self, base_program, logger, timeout=None,
                 weights=None, inputs=None, *args, **kwds):
        super(Exercise, self).__init__(*args, **kwds)

        logger.info('Setting up exercise with base_program %s' % base_program)

        if callable(base_program):
            self.base_program = FakeProgram(base_program, timeout)
            self.timeout = timeout 
        elif isinstance(base_program, Program):
            self.base_program = base_program
            self.timeout = base_program.timeout
        else:
            raise NotImplementedError("Never be here.")

        logger.info('Compiling the base_program.')
        s = self.base_program.compile(logger)
        if not s:
            raise ValueError("%s compile error" % self.base_program)

        self._set_up_weights(inputs, weights, logger)
        logger.info('Done setting up the exercise.')

    def _set_up_weights(self, inputs, weights, logger):
        if inputs is None:
            inputs = [None]
        if weights is None:
            weights = [20.0] + [80.0/len(inputs) for _ in inputs]
        if len(weights) != len(inputs) + 1:
            mesg = "Weights %s and inputs %s are inconsistent" % (weights, inputs)
            logger.error(mesg)
            raise ValueError(mesg)
        self.weights, self.inputs = weights, inputs

    def _check(self, inp, outp, base_outp, this_logger):
        """Compare the outputs given input, return the score out of 100.

        Subclass and override this method if you want a different way of
        of checking correctness (e.g. np.allclose, etc).

        """
        return 100 if outp.splitlines() == base_outp.splitlines() else 0

    def _parse_output(self, inp, outp, this_logger):
        """Coerce outp into the format suitable for comparison with base_outp.

        In case of failure returns None.

        Use case: `base_outp` is a list of floats, and `outp` is a string
        collected from stdout. Subclass and use this method to extract floats
        from the string.
        """
        return outp

    def _prepare_input(self, inp):
        """Prepare input for self.run (e.g. command line input)."""
        return str(inp)

    def mark(self, folder, logger, timeout=None):
        logger.info("*** Marking %s ***" % folder)

        if timeout is None:
            timeout = self.timeout

        try:
            submission = Submission(folder)
        except ValueError:
            # failed to find an executable
            logger.error("Failed to find an executable in %s." % folder)
            return 0

        program = Program(submission.folder, submission.fname, logger, timeout)

        with use_folder(submission.folder):
            # start marking
            mark = 0

            # compile
            success = program.compile(logger)
            if success:
                mark = self.weights[0]
            else:
                logger.info("Compilation failed, done marking. Mark = %s." % mark)
                return mark
            logger.info("Compilation success, mark = %s." % mark)
                    
            for inp, weight in zip(self.inputs, self.weights[1:]):
                logger.info("Checking input = %s" % inp)
                inp_ = self._prepare_input(inp)

                outp, err = program.run(logger, inp_)
                if err:
                    logger.error("stderr is \n===\n%s\n===\n" % err)
                    continue
                logger.info("Received output: %s." % outp)

                base_outp, base_err = self.base_program.run(logger, inp)
                if base_err and not err:
                    logger.error("base_stderr is %s " % base_err)
                    raise ValueError("base_err is %s for input %s " % (inp, base_err))

                # check/compare outp and base_outp
                result = 0
                outp_ = None
                try:
                    outp_ = self._parse_output(inp, outp, logger)
                except Exception as e:
                    result = 0
                    mesg = "Failed to parse the output: \n===\n %s\n===\n" % outp
                    mesg += "Exception: %s " % e
                    logger.error(mesg)

                if outp_:
                    try:     
                        result = self._check(inp, outp_, base_outp, logger)
                    except Exception as e:
                        result = 0
                        logger.error("Checking raised:  %s." % e)

                mark += result * weight / 100
                logger.info("result is %s, mark is %s out of %s." % (result,
                            mark, sum(self.weights)))
        logger.info("Done marking: %s out of %s" % (mark, sum(self.weights)))
        return mark

    def grade(self, *args, **kwds):
        """An alias for `mark`."""
        return self.mark(*args, **kwds)


def mark_one_path(mark_func, ppath, student, root_logger):

    # first of all, set up the per-student logger
    name = name_from_path(ppath)

    log_buf = StringIO()
    this_logger = setup_logger(logger_name=name,
                               log_file=os.path.join(ppath, name + '.log'),
                               level=logging.INFO,
                               log_buffer=log_buf)

    root_logger.info("Marking.. %s." % ppath)
    try:
        mark = mark_func(ppath, this_logger)
    except Exception as e:
        root_logger.error("Unknown exception: %s." % e)
        mark = 0

    root_logger.info("Done %s; mark = %s." % (ppath, mark))
    dct = {"name": student.name,
           "lms_id": name,
           "mark": mark,
           "log": log_buf.getvalue(),}
    log_buf.close()
    return dct


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("path",
                        help="Path to exercise folder or a single exercise.")
    parser.add_argument("--only", action="store_true",
                        help="Only mark a single exercise at path.")
    parser.add_argument("--checker", required=True,
                        help="The checker factory (required). Given X, the "
                              "factory is shims.get_X().")
    args = parser.parse_args()

    a_path = os.path.abspath(args.path)
    if not os.path.exists(a_path):
        raise ValueError("Path %s does not exist" % args.path)

    # figure out the root folder:
    #    if marking a single submission, root is os.getcwd()
    root_dir = args.path
    if args.only:
        root_dir = os.getcwd()

    # set up the root logger
    root_logger = setup_logger('root',
                               log_file=os.path.join(root_dir, 'root_log.log'))

    # select the exercise to mark
    #   XXX: some more flexibility: for now shims.py is hardcoded;
    #        a global registry of checkers? use --checker=module.factory CLI syntax? 
    shims = __import__('shims')
    factory = getattr(shims, 'get_'+args.checker)
    ex = factory(logger=root_logger)

    # Get the cohort: names, LMS ids etc
    cohort = fill_cohort()

    # Mark it
    if args.only:
        # assume the folder name is the lms_id
        lms_id = name_from_path(args.path)
        try:
            student = cohort[lms_id]
        except KeyError:
            student = Student(lms_id)
            cohort.update({lms_id: student})

        res = mark_one_path(ex.mark, args.path, student, root_logger)
        student.mark = res["mark"]
        results = [student]

    else:
        # walk: **Use abspaths, see os.walk docstring's last line**
        # root folder is path
        # The structure is 
        # root_dir
        #    - student_1
        #    - student_2
        #    - student_3
        # where each student_# is a directory with an executable. 
        results = []
        root_path, dirs, fnames = next(os.walk(root_dir))
        for folder in dirs:

            lms_id = folder   # assume this
            try:
                student = cohort[lms_id]
            except KeyError:
                student = Student(lms_id)

            ppath = os.path.join(root_path, folder)
            res = mark_one_path(ex.mark, ppath, student, root_logger)

            student.mark = round(res["mark"])
            student.log = res["log"]
            results.append(student)


        # now save results to Excel, for a good measure
        xls_path = os.path.join(root_path, "mark_result.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"], ws["B1"], ws["C1"] = "Name", "Mark", "Log"
        for row, student in enumerate(results):
            ws["A" + str(row+2)] = student.name
            ws["B" + str(row+2)] = student.mark
            ws["C" + str(row+2)] = student.log
        wb.save(xls_path)

    # print out the summary
    maxlen = max(len(_.name) for _ in results)
    fmt = "%"  + str(maxlen) + "s"
    print("\n\n", "*"*20, " Marks summary:")
    for entry in results:
        print(fmt % entry.name, " : ", entry.mark)
